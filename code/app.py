import os
import random
import string
from flask import Flask, render_template, request, redirect, url_for, flash, session
from openpyxl import Workbook, load_workbook
from email.mime.text import MIMEText
import smtplib

app = Flask(__name__)
app.secret_key = "my-secret-key"

GMAIL = "sagaunreal@gmail.com"
GMAIL_APP_PASSWORD = "qcmrrpxqauwqklij"  # Remove spaces for SMTP to work

USERS_XLSX = "users.xlsx"

# ---------------- Excel helpers ----------------
def ensure_workbook():
    if not os.path.exists(USERS_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = "users"
        ws.append(["UID", "Name", "Age", "Email", "Phone", "Password", "Verified"])
        wb.save(USERS_XLSX)

def read_all_users():
    ensure_workbook()
    wb = load_workbook(USERS_XLSX)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    users = []
    for r in rows:
        users.append({
            "uid": r[0],
            "name": r[1],
            "age": r[2],
            "email": r[3],
            "phone": r[4],
            "password": r[5],
            "verified": r[6],
        })
    return users

def store_user(uid, name, age, email, phone, password, verified=False):
    ensure_workbook()
    wb = load_workbook(USERS_XLSX)
    ws = wb.active
    ws.append([uid, name, age, email, phone, password, verified])
    wb.save(USERS_XLSX)

def user_exists(email=None, phone=None, uid=None):
    for u in read_all_users():
        if email and u["email"] == email:
            return True
        if phone and u["phone"] == phone:
            return True
        if uid and u["uid"] == uid:
            return True
    return False

def find_user_by_uid(uid):
    for u in read_all_users():
        if u["uid"] == uid:
            return u
    return None

# ---------------- UID generator ----------------
def generate_uid(n=8):
    while True:
        uid = ''.join(random.choices(string.digits, k=n))
        if not user_exists(uid=uid):
            return uid

# ---------------- Email sender ----------------
def send_email(to_email, subject, body):
    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = GMAIL
    msg["To"] = to_email

    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login(GMAIL, GMAIL_APP_PASSWORD)
    server.send_message(msg)
    server.quit()

# ---------------- Routes ----------------
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        uid = request.form.get("uid").strip()
        password = request.form.get("password").strip()

        user = find_user_by_uid(uid)
        if not user:
            flash("UID not found", "error")
            return redirect(url_for("login"))

        if user["password"] != password:
            flash("Wrong password", "error")
            return redirect(url_for("login"))

        if not user["verified"]:
            flash("Account not verified", "error")
            return redirect(url_for("login"))

        session["uid"] = uid
        return redirect(url_for("dashboard"))

    return render_template("login.html")

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form["name"].strip()
        age = request.form["age"].strip()
        email = request.form["email"].strip().lower()
        phone = request.form["phone"].strip()
        password = request.form["password"].strip()

        if user_exists(email=email):
            flash("Email already registered!", "error")
            return redirect(url_for("register"))

        if user_exists(phone=phone):
            flash("Phone already registered!", "error")
            return redirect(url_for("register"))

        otp = ''.join(random.choices(string.digits, k=6))

        session["pending_user"] = {
            "name": name,
            "age": age,
            "email": email,
            "phone": phone,
            "password": password
        }

        session["pending_otp"] = otp

        send_email(email, "Your OTP", f"Your OTP is: {otp}")

        flash("OTP sent to your email", "info")
        return redirect(url_for("verify_otp"))

    return render_template("register.html")

@app.route("/verify-otp", methods=["GET", "POST"])
def verify_otp():
    pending = session.get("pending_user")
    if not pending:
        flash("No pending registration.", "error")
        return redirect(url_for("register"))

    if request.method == "POST":
        entered = request.form["otp"].strip()
        real = session["pending_otp"]

        if entered == real:
            uid = generate_uid()
            store_user(uid, pending["name"], pending["age"], pending["email"],
                       pending["phone"], pending["password"], True)

            session.pop("pending_user")
            session.pop("pending_otp")

            flash(f"Registration complete! Your UID: {uid}", "success")
            return redirect(url_for("login"))
        else:
            flash("Wrong OTP!", "error")
            return redirect(url_for("verify_otp"))

    return render_template("verify_otp.html", email=pending["email"])

@app.route("/dashboard")
def dashboard():
    uid = session.get("uid")
    if not uid:
        return redirect(url_for("login"))
    user = find_user_by_uid(uid)
    return render_template("dashboard.html", user=user)

if __name__ == "__main__":
    ensure_workbook()
    app.run(debug=True)
