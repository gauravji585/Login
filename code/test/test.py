import os
import random
import string
from flask import Flask, render_template, request, redirect, url_for, flash, session
from openpyxl import Workbook, load_workbook
from email.mime.text import MIMEText
import smtplib




USERS_XLSX = "userse.xlsx"


def ensure_workbook():
    if not os.path.exists(USERS_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = "userse"
        ws.append(["UID", "Name", "Age", "Email", "Phone", "Password", "Verified"])
        wb.save(USERS_XLSX)
def read_all_users():
    ensure_workbook()
    wb = load_workbook(USERS_XLSX)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    users = []
    for r in rows:
        users.append({
            "uid": r[0],
            "name": r[1],
            "age": r[2],
            "email": r[3],
            "phone": r[4],
            "password": r[5],
            "verified": r[6],
        })
    return users

if __name__ == "__main__":
    ensure_workbook()